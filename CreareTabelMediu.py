import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
def apply_cell_colors(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row[1:]:
            if cell.value == 'Passed':
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
            elif cell.value == 'Failed':
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
            elif cell.value == 'Missing':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
            elif cell.value == 'Error':
                cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
def ConvToMediu(excel_file):
    df = pd.read_excel(excel_file, engine='openpyxl')
    dates = df.columns[1:]
    table = []
    for i in dates:
        date = i
        passed = df[i].value_counts().get('Passed', 0)
        failed = df[i].value_counts().get('Failed', 0)
        missing = df[i].value_counts().get('Missing', 0)
        table.append([date, passed, failed, missing])
    df = pd.DataFrame(data=table, columns=['Date', 'Passed', 'Failed', 'Missing'])
    df['Date'] = pd.to_datetime(df['Date'], dayfirst=True)
    df['Date'] = df['Date'].dt.date
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        # Write df2 to a new sheet named "Sheet2" (you can change the sheet name if needed)
        df.to_excel(writer, sheet_name='Test Results', index=False)
    book = openpyxl.load_workbook(excel_file)
    ws = book['Test Results']
    apply_cell_colors(ws)
    book.save(excel_file)
    book.close()
